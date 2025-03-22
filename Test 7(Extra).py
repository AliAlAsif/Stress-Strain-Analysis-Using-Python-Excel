import openpyxl
import matplotlib.pyplot as plt
import numpy as np
from openpyxl.drawing.image import Image
from scipy.stats import linregress
from scipy.integrate import simpson

# File paths and constants
FILE_PATH = 'For test.xlsx'
UPDATED_FILE_PATH = 'updated_' + FILE_PATH
SEARCH_WORDS = ['Force', 'Length, L', 'Width, W', 'Stroke', 'Thickness, T', 'Maximum Stress, σc']


def load_excel(file_path):
    """Load an Excel workbook."""
    try:
        return openpyxl.load_workbook(file_path)
    except Exception as e:
        raise FileNotFoundError(f"Failed to load file {file_path}: {e}")


def find_words_in_excel(sheet, search_words):
    """Find positions of search words in an Excel sheet."""
    positions = {word: [] for word in search_words}
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value in search_words:
                positions[cell.value].append((cell.row, cell.column))
    return positions


def find_second_cell_value(sheet, positions, word):
    """Retrieve the numeric value two rows below the specified word."""
    pos_list = positions.get(word, [])
    if not pos_list:
        raise ValueError(f"'{word}' not found in the Excel file.")
    row, col = pos_list[0]
    value = sheet.cell(row=row + 2, column=col).value
    if value is None or not isinstance(value, (int, float)):
        raise ValueError(f"The cell two rows below '{word}' does not contain a numeric value.")
    return float(value)


def calculate_stress_and_strain(file_path, search_words):
    """Calculate and save stress and strain values in the Excel sheet."""
    workbook = load_excel(file_path)
    sheet = workbook.active
    positions = find_words_in_excel(sheet, search_words + ['Stress', 'Strain'])

    # Extract required values
    length = find_second_cell_value(sheet, positions, 'Length, L')
    width = find_second_cell_value(sheet, positions, 'Width, W')
    thickness = find_second_cell_value(sheet, positions, 'Thickness, T')

    # Calculate stress
    stress = []
    for force_row, force_col in positions.get('Force', []):
        for r in range(force_row + 1, sheet.max_row + 1):
            force = sheet.cell(row=r, column=force_col).value
            if isinstance(force, (int, float)):
                stress.append(force / (length * width))

    # Save stress
    if 'Stress' in positions:
        stress_row, stress_col = positions['Stress'][0]
        for i, s in enumerate(stress):
            sheet.cell(row=stress_row + i + 2, column=stress_col, value=s)

    # Calculate strain
    strain = []
    for stroke_row, stroke_col in positions.get('Stroke', []):
        for r in range(stroke_row + 1, sheet.max_row + 1):
            stroke = sheet.cell(row=r, column=stroke_col).value
            if isinstance(stroke, (int, float)):
                strain.append((stroke / thickness) * 100)

    # Save strain
    if 'Strain' in positions:
        strain_row, strain_col = positions['Strain'][0]
        for i, st in enumerate(strain):
            sheet.cell(row=strain_row + i + 2, column=strain_col, value=st)

    workbook.save(UPDATED_FILE_PATH)
    return stress, strain, UPDATED_FILE_PATH


def plot_stress_strain_curve(stress, strain, file_path):
    """Plot and save the stress-strain curve to the Excel file."""
    if len(stress) != len(strain):
        raise ValueError("Stress and strain lengths mismatch.")

    # Plot
    plt.figure(figsize=(8, 6))
    plt.plot(strain, stress, marker='o', label='Stress-Strain Curve')
    plt.xlabel('Strain (%)')
    plt.ylabel('Stress (MPa)')
    plt.title('Stress-Strain Curve')
    plt.grid(True)
    plt.legend()
    plot_file = 'stress_strain_plot.png'
    plt.savefig(plot_file)
    plt.close()

    # Insert into Excel
    workbook = load_excel(file_path)
    sheet = workbook.active
    img = Image(plot_file)
    img.width, img.height = 400, 300
    sheet.add_image(img, 'A20')
    workbook.save(file_path)


def calculate_compression_modulus(stress, strain, file_path):
    """Calculate and save the Compression Modulus (Ec) in the Excel file."""
    stress = np.array(stress)
    strain = np.array(strain)

    window_size = 50
    max_r2, best_slope = -1, 0

    for i in range(len(strain) - window_size + 1):
        subset_strain = strain[i:i + window_size]
        subset_stress = stress[i:i + window_size]
        slope, _, r2, _, _ = linregress(subset_strain, subset_stress)
        if r2 > max_r2:
            max_r2, best_slope = r2, slope

    workbook = load_excel(file_path)
    sheet = workbook.active
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == 'Comp. Modulus, Ec':
                sheet.cell(row=cell.row + 2, column=cell.column, value=best_slope)
    workbook.save(file_path)


def calculate_maximum_stress(sheet, positions, file_path):
    """Find and save the maximum stress in the Excel sheet."""
    if 'Stress' not in positions or 'Maximum Stress, σc' not in positions:
        raise ValueError("'Stress' or 'Maximum Stress, σc' not found in the Excel file.")

    # Get stress column
    stress_col = positions['Stress'][0][1]
    stress_values = [
        sheet.cell(row=r, column=stress_col).value
        for r in range(2, sheet.max_row + 1)  # Start from row 2 to skip headers
        if isinstance(sheet.cell(row=r, column=stress_col).value, (int, float))
    ]

    if not stress_values:
        raise ValueError("No numeric values found in the 'Stress' column.")

    # Find the maximum stress
    max_stress = max(stress_values)
    print(f"Calculated Maximum Stress: {max_stress} MPa")

    # Get the cell position for "Maximum Stress, σc"
    max_stress_row, max_stress_col = positions['Maximum Stress, σc'][0]

    # Save the value in the sheet
    sheet.cell(row=max_stress_row + 2, column=max_stress_col, value=max_stress)

    # Save the workbook
    workbook = sheet.parent
    workbook.save(file_path)
    print("Workbook saved successfully.")


def calculate_energy_upto_strain(stress, strain, file_path):
    """Calculate and save the energy under the stress-strain curve up to a user-specified strain percentage."""
    stress = np.array(stress)
    strain = np.array(strain)

    # Prompt user for maximum strain percentage
    max_strain_percentage = float(input("Enter the maximum strain percentage: "))  # Get percentage from user
    max_strain = max_strain_percentage / 100  # Convert percentage to a decimal value

    # Filter the stress and strain arrays based on the chosen strain percentage
    valid_indices = strain <= max_strain
    strain_subset = strain[valid_indices]
    stress_subset = stress[valid_indices]

    if len(strain_subset) < 2:
        raise ValueError(f"Not enough data points below {max_strain_percentage}% strain.")

    # Calculate energy using Simpson's rule
    energy = simpson(stress_subset, x=strain_subset)

    # Generate the dynamic label based on the user input
    energy_label = f"Energy up to {max_strain_percentage}% Strain, E0.{str(int(max_strain_percentage))}"

    # Load the workbook and search for the label
    workbook = load_excel(file_path)
    sheet = workbook.active

    # Search for the label and update the value
    label_found = False
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and energy_label in str(cell.value):
                # Save the calculated energy under the correct label
                sheet.cell(row=cell.row + 2, column=cell.column, value=energy)
                label_found = True
                break
        if label_found:
            break

    if not label_found:
        raise ValueError(f"Label '{energy_label}' not found in the Excel sheet.")

    # Save the workbook
    workbook.save(file_path)
    print(f"Energy up to {max_strain_percentage}% strain saved: {energy} MPa*%")



# Main execution
try:
    stress, strain, updated_file = calculate_stress_and_strain(FILE_PATH, SEARCH_WORDS)
    workbook = load_excel(updated_file)
    sheet = workbook.active
    positions = find_words_in_excel(sheet,
                                    SEARCH_WORDS + ['Stress', 'Strain', 'Maximum Stress, σc', 'Comp. Modulus, Ec',
                                                    'Energy up to 40.0% Strain, E0.40'])

    calculate_maximum_stress(sheet, positions, updated_file)
    calculate_compression_modulus(stress, strain, updated_file)
    calculate_energy_upto_strain(stress, strain, updated_file)
    plot_stress_strain_curve(stress, strain, updated_file)
except Exception as e:
    print(f"Error: {e}")
