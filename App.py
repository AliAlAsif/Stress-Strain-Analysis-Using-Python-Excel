import openpyxl
import matplotlib.pyplot as plt
import numpy as np
from openpyxl.drawing.image import Image
from scipy.stats import linregress
from scipy.integrate import simpson
import tkinter as tk
from tkinter import filedialog, messagebox
import os

# File paths and constants
SEARCH_WORDS = ['Force', 'Length, L', 'Width, W', 'Stroke', 'Thickness, T', 'Maximum Stress, σc']


def load_excel(file_path):
    """Load an Excel workbook."""
    try:
        return openpyxl.load_workbook(file_path)
    except Exception as e:
        raise FileNotFoundError(f"Failed to load file {file_path}: {e}")


def find_words_in_excel(sheet, search_words):
    """Find positions of search words in an Excel sheet."""
    positions = {word: [] for word in search_words}
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value in search_words:
                positions[cell.value].append((cell.row, cell.column))
    return positions


def find_second_cell_value(sheet, positions, word):
    """Retrieve the numeric value two rows below the specified word."""
    pos_list = positions.get(word, [])
    if not pos_list:
        raise ValueError(f"'{word}' not found in the Excel file.")
    row, col = pos_list[0]
    value = sheet.cell(row=row + 2, column=col).value
    if value is None or not isinstance(value, (int, float)):
        raise ValueError(f"The cell two rows below '{word}' does not contain a numeric value.")
    return float(value)


def calculate_stress_and_strain(file_path, search_words):
    """Calculate and save stress and strain values in the Excel sheet."""
    workbook = load_excel(file_path)
    sheet = workbook.active
    positions = find_words_in_excel(sheet, search_words + ['Stress', 'Strain'])

    # Extract required values
    length = find_second_cell_value(sheet, positions, 'Length, L')
    width = find_second_cell_value(sheet, positions, 'Width, W')
    thickness = find_second_cell_value(sheet, positions, 'Thickness, T')

    # Calculate stress
    stress = []
    for force_row, force_col in positions.get('Force', []):
        for r in range(force_row + 1, sheet.max_row + 1):
            force = sheet.cell(row=r, column=force_col).value
            if isinstance(force, (int, float)):
                stress.append(force / (length * width))

    # Save stress
    if 'Stress' in positions:
        stress_row, stress_col = positions['Stress'][0]
        for i, s in enumerate(stress):
            sheet.cell(row=stress_row + i + 2, column=stress_col, value=s)

    # Calculate strain
    strain = []
    for stroke_row, stroke_col in positions.get('Stroke', []):
        for r in range(stroke_row + 1, sheet.max_row + 1):
            stroke = sheet.cell(row=r, column=stroke_col).value
            if isinstance(stroke, (int, float)):
                strain.append((stroke / thickness) * 100)

    # Save strain
    if 'Strain' in positions:
        strain_row, strain_col = positions['Strain'][0]
        for i, st in enumerate(strain):
            sheet.cell(row=strain_row + i + 2, column=strain_col, value=st)

    # Save the updated file in the same directory as the input file
    directory = os.path.dirname(file_path)  # Get the directory of the input file
    updated_file_name = 'updated_' + os.path.basename(file_path)  # Create the updated file name
    updated_file_path = os.path.join(directory, updated_file_name)  # Full path for the updated file

    workbook.save(updated_file_path)
    return stress, strain, updated_file_path


def plot_stress_strain_curve(stress, strain, file_path):
    """Plot and save the stress-strain curve to the Excel file."""
    if len(stress) != len(strain):
        raise ValueError("Stress and strain lengths mismatch.")

    # Plot
    plt.figure(figsize=(8, 6))
    plt.plot(strain, stress, marker='o', label='Stress-Strain Curve')
    plt.xlabel('Strain (%)')
    plt.ylabel('Stress (MPa)')
    plt.title('Stress-Strain Curve')
    plt.grid(True)
    plt.legend()
    plot_file = 'stress_strain_plot.png'
    plt.savefig(plot_file)
    plt.close()

    # Insert into Excel
    workbook = load_excel(file_path)
    sheet = workbook.active
    img = Image(plot_file)
    img.width, img.height = 400, 300
    sheet.add_image(img, 'A20')
    workbook.save(file_path)


def calculate_compression_modulus(stress, strain, file_path):
    """Calculate and save the Compression Modulus (Ec) in the Excel file."""
    stress = np.array(stress)
    strain = np.array(strain)

    window_size = 50
    max_r2, best_slope = -1, 0

    for i in range(len(strain) - window_size + 1):
        subset_strain = strain[i:i + window_size]
        subset_stress = stress[i:i + window_size]
        slope, _, r2, _, _ = linregress(subset_strain, subset_stress)
        if r2 > max_r2:
            max_r2, best_slope = r2, slope

    workbook = load_excel(file_path)
    sheet = workbook.active
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == 'Comp. Modulus, Ec':
                sheet.cell(row=cell.row + 2, column=cell.column, value=best_slope)
    workbook.save(file_path)


def calculate_maximum_stress(sheet, positions, file_path):
    """Find and save the maximum stress in the Excel sheet."""
    if 'Stress' not in positions or 'Maximum Stress, σc' not in positions:
        raise ValueError("'Stress' or 'Maximum Stress, σc' not found in the Excel file.")

    # Get stress column
    stress_col = positions['Stress'][0][1]
    stress_values = [
        sheet.cell(row=r, column=stress_col).value
        for r in range(2, sheet.max_row + 1)  # Start from row 2 to skip headers
        if isinstance(sheet.cell(row=r, column=stress_col).value, (int, float))
    ]

    if not stress_values:
        raise ValueError("No numeric values found in the 'Stress' column.")

    # Find the maximum stress
    max_stress = max(stress_values)
    print(f"Calculated Maximum Stress: {max_stress} MPa")

    # Get the cell position for "Maximum Stress, σc"
    max_stress_row, max_stress_col = positions['Maximum Stress, σc'][0]

    # Save the value in the sheet
    sheet.cell(row=max_stress_row + 2, column=max_stress_col, value=max_stress)

    # Save the workbook
    workbook = sheet.parent
    workbook.save(file_path)
    print("Workbook saved successfully.")


def calculate_energy_upto_strain(stress, strain, file_path):
    """Calculate and save the energy under the stress-strain curve up to 40% strain."""
    stress = np.array(stress)
    strain = np.array(strain)

    # Fixed maximum strain percentage
    max_strain_percentage = 40  # Hardcoded to 40%
    max_strain = max_strain_percentage / 100  # Convert percentage to a decimal value

    # Filter the stress and strain arrays based on the chosen strain percentage
    valid_indices = strain <= max_strain
    strain_subset = strain[valid_indices]
    stress_subset = stress[valid_indices]

    if len(strain_subset) < 2:
        raise ValueError(f"Not enough data points below {max_strain_percentage}% strain.")

    # Calculate energy using Simpson's rule
    energy = simpson(stress_subset, x=strain_subset)

    # Generate the fixed label
    energy_label = f"Energy up to {max_strain_percentage}% Strain, E0.4"

    # Load the workbook and search for the label
    workbook = load_excel(file_path)
    sheet = workbook.active

    # Search for the label and update the value
    label_found = False
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and energy_label in str(cell.value):
                # Save the calculated energy under the correct label
                sheet.cell(row=cell.row + 2, column=cell.column, value=energy)
                label_found = True
                break
        if label_found:
            break

    if not label_found:
        raise ValueError(f"Label '{energy_label}' not found in the Excel sheet.")

    # Save the workbook
    workbook.save(file_path)
    print(f"Energy up to {max_strain_percentage}% strain saved: {energy} MPa*%")


class CompressionTestApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Compression Test Analysis")
        self.file_path = None
        self.stress = None
        self.strain = None
        self.updated_file = None

        # Create GUI elements
        self.label = tk.Label(root, text="Compression Test Analysis", font=("Arial", 16))
        self.label.pack(pady=10)

        self.load_button = tk.Button(root, text="Load Excel File", command=self.load_file)
        self.load_button.pack(pady=5)

        self.calculate_button = tk.Button(root, text="Calculate Stress and Strain", command=self.calculate_stress_strain)
        self.calculate_button.pack(pady=5)

        self.max_stress_button = tk.Button(root, text="Calculate Maximum Stress", command=self.calculate_max_stress)
        self.max_stress_button.pack(pady=5)

        self.modulus_button = tk.Button(root, text="Calculate Compression Modulus", command=self.calculate_modulus)
        self.modulus_button.pack(pady=5)

        self.energy_button = tk.Button(root, text="Calculate Energy up to 40% Strain", command=self.calculate_energy)
        self.energy_button.pack(pady=5)

        self.plot_button = tk.Button(root, text="Plot Stress-Strain Curve", command=self.plot_curve)
        self.plot_button.pack(pady=5)

    def load_file(self):
        """Load the Excel file."""
        self.file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if self.file_path:
            messagebox.showinfo("File Loaded", f"File {self.file_path} loaded successfully.")

    def calculate_stress_strain(self):
        """Calculate stress and strain."""
        if not self.file_path:
            messagebox.showerror("Error", "Please load an Excel file first.")
            return

        try:
            self.stress, self.strain, self.updated_file = calculate_stress_and_strain(self.file_path, SEARCH_WORDS)
            messagebox.showinfo("Success", f"Stress and strain calculated and saved to:\n{self.updated_file}")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def calculate_max_stress(self):
        """Calculate maximum stress."""
        if not self.updated_file:
            messagebox.showerror("Error", "Please calculate stress and strain first.")
            return

        try:
            workbook = load_excel(self.updated_file)
            sheet = workbook.active
            positions = find_words_in_excel(sheet, SEARCH_WORDS + ['Stress', 'Maximum Stress, σc'])
            calculate_maximum_stress(sheet, positions, self.updated_file)
            messagebox.showinfo("Success", "Maximum stress calculated and saved successfully.")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def calculate_modulus(self):
        """Calculate compression modulus."""
        if not self.stress or not self.strain:
            messagebox.showerror("Error", "Please calculate stress and strain first.")
            return

        try:
            calculate_compression_modulus(self.stress, self.strain, self.updated_file)
            messagebox.showinfo("Success", "Compression modulus calculated and saved successfully.")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def calculate_energy(self):
        """Calculate energy up to 40% strain."""
        if not self.stress or not self.strain:
            messagebox.showerror("Error", "Please calculate stress and strain first.")
            return

        try:
            calculate_energy_upto_strain(self.stress, self.strain, self.updated_file)
            messagebox.showinfo("Success", "Energy up to 40% strain calculated and saved.")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def plot_curve(self):
        """Plot the stress-strain curve."""
        if not self.stress or not self.strain:
            messagebox.showerror("Error", "Please calculate stress and strain first.")
            return

        try:
            plot_stress_strain_curve(self.stress, self.strain, self.updated_file)
            messagebox.showinfo("Success", "Stress-strain curve plotted and saved successfully.")
        except Exception as e:
            messagebox.showerror("Error", str(e))


# Run the application
if __name__ == "__main__":
    root = tk.Tk()
    app = CompressionTestApp(root)
    root.mainloop()